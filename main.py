import flet as ft
import datetime, openpyxl, os, random
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.window import WindowTypes
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl.styles import PatternFill, Font
from settings import prev_proxies, proxies, default_proxy, global_url

driver = None
dd = None
home = None
away = None
h1 = None
dX = None
w2 = None
p1 = None
p2 = None
p3 = None
p4 = None
score1 = None
score2 = None
today = datetime.date.today()
selection = "All"

def calculate_abs(cellvalue):
    try:
        numbers = cellvalue.split('-')
        num1 = int(numbers[0].strip())
        num2 = int(numbers[1].strip())  
        result = abs(num1 - num2)
        return result
    except:
        return 0


def get_analysis(e):
    
    # Get the total number of rows with content
    total_rows = sheet.max_row

    print("Total rows with content:", total_rows)
    
    for row in range(total_rows):
  
        if row == 0:
            print("skip")
            
        else:
            cellK = sheet["K" + str(row + 1)]
            if str(cellK.value).lower() == "w":
                cellK.fill = PatternFill("solid", fgColor="009999FF")
                cellK.font = Font(bold=True, color="00FFFFFF")
            
            
            cellN = sheet["N" + str(row + 1)]
            if str(cellN.value).lower() == "w":
                cellN.fill = PatternFill("solid", fgColor="00FFCC00")
                cellK.font = Font(bold=True, color="00333333")
            
            
            cellJ = sheet["J" + str(row + 1)]
            cellJJ = calculate_abs(cellJ.value)
    
            if cellJJ > 2:
                cellJ.fill = PatternFill("solid", fgColor="00008000")
                cellJ.font = Font(bold=True, color="00FFFFFF")
                
            
            cellM = sheet["M" + str(row + 1)]
            cellMM = calculate_abs(cellM.value)

            if cellMM > 3:
                cellM.fill = PatternFill("solid", fgColor="00FF6600")
                cellM.font = Font(bold=True, color="00FFFFFF")
                
                

            if (cellJJ > 2) and (str(cellK.value).lower() == "w"):
                row1 = sheet[row + 1]
                
                for index, r in enumerate(row1):
                    r.fill = PatternFill("solid", fgColor="00CCFFFF")
                    if index > len(sheet['A:V']):
                        break
 
                
            if (cellJJ > 2) and (str(cellK.value).lower() == "w"):
                row1 = sheet[row + 1]
                
                for index, r in enumerate(row1):
                    r.fill = PatternFill("solid", fgColor="00CCFFFF")
                    if index > len(sheet['A:V']):
                        break
                

    
    workbook.save(file_path)
    os.startfile(file_path)

    
### putting the score to the excel file.. back..
def fill_score(all_scores):
    last_row = len(all_scores)
    
    font = openpyxl.styles.Font(bold=True, color="FF0000")
    blue = openpyxl.styles.Font(bold=True, color="00FF00")
    
    for row in range(0, last_row + 1):
    
        if row == 0:
            cellV = sheet["V" + str(row + 1)]
            cellV.value = "Scores"
            cellV.font = font
        else:
            sheet["V" + str(row + 1)] = all_scores[row - 1]
            
            cellT = sheet["T" + str(row + 1)]
            if str(cellT.value).lower() == "true":
                cellT.font = blue
                
            cellU = sheet["U" + str(row + 1)]
            if str(cellU.value).lower() == "true":
                cellU.font = blue
                
                
            
    
    workbook.save(file_path)
    os.startfile(file_path)

def get_all_scores(e):
    global driver
    
    scores_urls = column_b_url()
    all_scores = []
    
    if driver:

        for score_url in scores_urls:
            sleep(3)
            print(score_url)
            score = open_url_score(score_url)
            all_scores.append(score)
        
        fill_score(all_scores)

def get_scores(e):
    global driver
    
    try:  
        if driver:
            driver.switch_to.window(driver.window_handles[0])
        else:
            firefox_launch(e)
    except:
        pass
    
def open_url_score(url):
    global driver
    if driver:
        driver.get(url)
        
        driver.implicitly_wait(40)
        
        score = None
        sleep(2)
        scores_tag = driver.find_element(By.XPATH, "//div[@class='duelParticipant__score']//div[@class='detailScore__matchInfo']")
                    
        score_text = scores_tag.get_attribute("innerText")
        score_list = score_text.splitlines()
        
        try:
            HScore = score_list[0]
            AScore = score_list[2]
            score = f"{HScore} - {AScore}"
        except:
            score = ""
            pass
        
        return score

def column_b_url():
    column_letter = "B"
    scores_urls = []
    urls = [cell.value for cell in sheet[column_letter]]
    for url in urls[1:]:
        scores_urls.append(url)
    
    return scores_urls

def open_file(e: ft.FilePickerResultEvent):
    global file_path, sheet, workbook
    file_path = ""
    sheet = None
    workbook = None
    selected_files.value = (
            ", ".join(map(lambda f: f.name, e.files)) if e.files else "Cancelled!"
        )
    selected_files.update()
    file_path  = ("".join(map(lambda f: f.path, e.files)) if e.files else None)
    
    if file_path is not None:
        process_xlsx_file(file_path)
              
def process_xlsx_file(file_path):
    global sheet, workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

def auto_run_data(e):
    global driver
    
    if driver:
        
        auto_run_btn.disabled = True
        auto_run_btn.update()
        
        all_urls = get_all_matches_url(e)
        
        open_new_tab(e)
        
        data = url_process(e, all_urls)
        
        save_file(data=data)

        sleep(2)
        
        # firefox_close(e)
        
        sleep(2)
        
        auto_open()

def url_process(e, all_urls):
    
    global driver, teams, home, away, p1, p2, p3, p4, h1, dX, w2, home_last_match, away_last_match, score1, score2
    
    data = []
    
    for index, url in enumerate(all_urls):
            
        if active:
            
            open_url(url)
            get_odds(e)
            sleep(3)
            get_last_match(e)
            sleep(3)
            try:
                gltH = get_last_team_home()
                gltW = get_last_team_away()
            except Exception as e:
                print(f"Error fro getting last team") 
                pass 
            try:
                teams = [home, away, gltH[0], gltW[0]]
            except:
                print("Error getting teams")
                pass
            sleep(3)
            details_from_table(e, teams)
    
            my_Hypo = my_hypothesis()
            my_Hypo2 = my_hypothesis2()
            
            try:
                data.append({
                    "index": index + 1,
                    "url": url,
                    "home": home,
                    "away": away,
                    "1": h1,
                    "x": dX,
                    "2": w2,
                    "HLM": home_last_match,
                    "gltH": gltH,
                    "Hscore": score1,
                    "Hsign": symbol1,
                    "ALM": away_last_match,
                    "Ascore": score2,
                    "Wsign": symbol2,
                    "gltW": gltW,
                    "p1": p1,
                    "p2": p2,
                    "p3": p3,
                    "p4": p4,
                    "Hypothesis": my_Hypo,
                    "Hypothesis2": my_Hypo2,
                })
            except:
                print("Error fixing table")
                
            # Reset
            
            try:
                reset()
            except Exception as e:
                print(f"Error resting values: {e}")
            
        else:
            print("Stopped..")
            break
        
    return data
                
def auto_open():
    directory_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output_data")  
    
    try:
        files = os.listdir(directory_path)
        
        file_info = [(file, os.path.getmtime(os.path.join(directory_path, file))) for file in files]

        file_info.sort(key=lambda x: x[1], reverse=True)


        sorted_files = [file_info[0][0]] 
        for file, _ in file_info[1:]:  
            sorted_files.append(file)
        
        if sorted_files:
            latest_file = sorted_files[0]
            file_path = os.path.join(directory_path, latest_file)
            print(file_path)    
            os.startfile(file_path)
        else:
            print("No files found in 'output_data' directory.")
            
    except FileNotFoundError:
        print(f"Directory '{directory_path}' not found. Creating it.")
        os.makedirs(directory_path, exist_ok=True)
        files = []
        
def save_file(data):
    
    today = datetime.date.today()
    filename = f"{selection}_{today.strftime('%Y-%m-%d')}.xlsx"
    try:
        directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output_data")
    except FileNotFoundError:
         print(f"Directory '{directory}' not found. Creating it.")
         os.makedirs(directory, exist_ok=True)

    os.makedirs(directory, exist_ok=True)
    file_path = os.path.join(directory, filename)
    df = pd.DataFrame(data=data)
    df.to_excel(file_path, index=False)
    
def my_hypothesis():
    # global p4, p1, symbol2
    try:
        if (p4, p1, symbol2 ):
            if(p4 < p1) and ((symbol2 == "D" or symbol2 == "L")):
                return True
            else:
                return False
    except Exception as e:
        print(f"Error getting hypo")
        return None
    
def my_hypothesis2():
    # global p4, p1, symbol2, p2
    try:
        if (p4, p1, symbol2,  ):
            if(p4 < p1) and ((symbol2 == "D" or symbol2 == "L") and (p1 < p2 )):
                return True
            else:
                return False
    except Exception as e:
        print(f"Error getting hypo")
        return None
    



def reset():
    global p1, p2, p3, p4, h1, dX, w2, home, away, score1, score2

    if p1 is not None:
        p1 = None
    if p2 is not None:
        p2 = None
    if p3 is not None:
        p3 = None
    if p4 is not None:
        p4 = None
    if h1 is not None:
        h1 = None
    if dX is not None:
        dX = None
    if w2 is not None:
        w2 = None
    if home is not None:
        home = None
    if away is not None:
        away = None
    if score1 is not None:
        score1 = None
    if score2 is not None:
        score2 = None

def details_from_table(e, teams):
    global driver, p1, p2, p3, p4
    
    if driver:
        try:
            driver.implicitly_wait(10)
            standing = driver.find_element(By.XPATH, "//button[normalize-space()='Standings']")
            standing.click()
            
            
            driver.implicitly_wait(30)
            total_matches = driver.find_elements(By.CLASS_NAME, "tableCellParticipant__name")
            
            print(len(total_matches))
            
            for index, matches in enumerate(total_matches):
                name = matches.get_attribute("innerText")
                # print(f"team {index + 1}: {name}")
                # if name in list
                if name in teams[0]:
                    print(f"Team founds position {index + 1} : {name}")
                    p1 = index + 1
                elif name in teams[1]:
                    print(f"Team founds position {index + 1} : {name}")
                    p2 = index + 1
                elif name in teams[2]:
                    print(f"Team founds position {index + 1} : {name}")
                    p3 = index + 1
                elif name in teams[3]:
                    print(f"Team founds position {index + 1} : {name}")
                    p4 = index + 1
                            
            
        except Exception as e:
            print(f"Error clicking on standing table")
            pass

def get_last_team_home():

    print("print the last team name..")
            
    if home_last_match[0] == home:
        # print("Team is Home")
        return (home_last_match[1], "1TIH")
    else:
        # print("Team is away")
        return (home_last_match[0], "1TIW")

def get_last_team_away():

    # print("print the last team name..")
    if away_last_match[0] == away:
        # print("Team is Home")
        return (away_last_match[1], "2TIH")
    else:
        # print("Team is away")
        return (away_last_match[0], "2TIW")

def rotate_proxy():
    global prev_proxies
    total_no = len(proxies)
    
    if not prev_proxies:
        rand = random.choice(proxies)
        return rand
    
    else: 
        for i, proxy in enumerate(proxies): 
            if not proxy in prev_proxies and len(prev_proxies) < total_no:
                prev_proxies.append(proxy)
                print(f"next proxy no: {i}")
                return proxy
            
            if len(prev_proxies) == total_no:
                print("clear proxy_list")
                prev_proxies.clear()
                prev_proxies.append(default_proxy)
                return default_proxy

def get_last_match(e):
    global driver, home_last_match, away_last_match, score1, symbol1, score2, symbol2
    
    if driver:
        try:
            driver.implicitly_wait(60)
            
            h2h = driver.find_element(By.XPATH, "//button[normalize-space()='H2H']")
            h2h.click()
            
            #logic...
            home_last_game_tag = driver.find_element(By.XPATH, "//body/div[@class='container__detail']/div[@id='detail']/div[@class='h2hSection']/div[@class='h2h']/div[1]/div[2]/div[1]")
            
            home_last_game = home_last_game_tag.get_attribute("innerText")
            home_last = home_last_game.splitlines()
            
            away_last_game_tag = driver.find_element(By.XPATH, "//body/div[@class='container__detail']/div[@id='detail']/div[@class='h2hSection']/div[@class='h2h']/div[2]/div[2]/div[1]")
            
            away_last_game =away_last_game_tag.get_attribute("innerText")
            away_last = away_last_game.splitlines()
            
            teamA = home_last[2]
            teamB = home_last[3]
            score1 = f"{home_last[4]} - {home_last[5]}"
            symbol1 = home_last[6]
            teamC = away_last[2]
            teamD = away_last[3]
            score2 = f"{away_last[4]} - {away_last[5]}"
            symbol2 = away_last[6]
            
            home_last_match = [teamA, teamB]
            away_last_match = [teamC, teamD]
        except:
            print("Error getting last_match")
            pass
    
def get_odds(e):
    global driver, h1, dX, w2
    
    if driver:
        driver.implicitly_wait(4)
        odds_tag = WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div[class='oddsRowContent']:first-of-type")))
        
        odds_text = odds_tag.get_attribute("innerText")
        odds_list = odds_text.splitlines()

        h1 = odds_list[1]
        dX = odds_list[3]
        w2 = odds_list[5]
     
def open_url(url):
    global home, away, driver
    
    if driver:
        try:
            if url:
                driver.get(url)
                
                driver.implicitly_wait(4)
                
                home_tag = driver.find_element(By.CSS_SELECTOR, "div[class='duelParticipant__home '] a[class='participant__participantName participant__overflow ']")
                home = home_tag.get_attribute("innerText")
                print(f"home: {home}")
                
                away_tag = driver.find_element(By.CSS_SELECTOR, "div[class='duelParticipant__away '] a[class='participant__participantName participant__overflow ']")
                away = away_tag.get_attribute("innerText")
                print(f"away: {away}")
                
                driver.implicitly_wait(4)
        except Exception as e:
            print("Url Error")
            pass

def open_new_tab(e):
    global driver
    
    if driver:
        driver.switch_to.new_window(WindowTypes.TAB)
        
        driver.switch_to.window(driver.window_handles[1])
        
        # open_url()
        
def get_all_matches_url(e):
    global driver, active
    
    all_matches_url_list = []
    
    if driver:
        elements = driver.find_elements(By.CLASS_NAME, "eventRowLink")
        active = True
        for element in elements:
            href = element.get_attribute("href")
            all_matches_url_list.append(href)
        
        return all_matches_url_list

def firefox_launch(e):
    global driver
    
    try:
        open.disabled = True
        open.update()
        firefox_options = Options()
        firefox_options.add_argument("--disable-infobars")
        firefox_options.add_argument("--disable-extensions")
        firefox_options.add_argument("--disable-popup-blocking")
        
        
        # options = webdriver.FirefoxOptions()
        proxy = rotate_proxy()
        print(proxy)
        firefox_options.add_argument(f"proxy-server={proxy}")
        # options.profile = webdriver.FirefoxProfile(pf)
        
        driver = webdriver.Firefox(options=firefox_options,)
        driver.maximize_window()
        driver.get(url=global_url)
        
        try:
            wait = WebDriverWait(driver, 60)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id='onetrust-reject-all-handler']")))
            element.click()
            driver.find_element(By.XPATH, "//button[@id='onetrust-reject-all-handler']").click()
        except:
            pass
        
    except:
        print("Error opening browser")

def odds_click(e):
    global driver
    
    if driver:
        driver.find_element(By.XPATH, "//div[contains(text(),'Odds')]").click()

def away_filter_finish(e):
    global driver, selection
    
    if driver:
        if awaydd.value == "Away 2.0 - 2.5":
            print(awaydd.value)
            selection = awaydd.value + " Finish"
            driver.execute_script("""matches = document.getElementsByClassName("event__match event__match--twoLine");
    for(i=matches.length -1; i>=0; i--){
        match = matches[i];
        if(match.children.length > 9){
            if(match.children[9] && match.children[10] && match.children[11]){
                odd_1 = parseFloat(match.children[9].textContent);
                odd_x = parseFloat(match.children[10].textContent);
                odd_2 = parseFloat(match.children[11].textContent);

                if((odd_1 >= 2.50 && odd_1 <= 8.0) && (odd_x >= 2.5 && odd_x <= 8.0) && (odd_2 >= 2.0 && odd_2 < 2.5)){
                    // keep code
                } else {
                    match.parentElement.removeChild(match);
                }

            } else {
                match.parentElement.removeChild(match);
            }
        }
    }""")
            
        elif awaydd.value == "Away 1.5 - 2.0":
            print(awaydd.value)
            selection = awaydd.value + " Finish"
            driver.execute_script("""matches = document.getElementsByClassName("event__match event__match--twoLine");
    for(i=matches.length-1; i>=0; i--){
        match = matches[i];
        if(match.children.length > 9){
            if(match.children[9] && match.children[10] && match.children[11]){
                odd_1 = parseFloat(match.children[9].textContent);
                odd_x = parseFloat(match.children[10].textContent);
                odd_2 = parseFloat(match.children[11].textContent);

                if((odd_1 >= 3.5 && odd_1 <= 8.0) && (odd_x >= 2.5 && odd_x <= 5.0) && (odd_2 >= 1.5 && odd_2 <= 2.0)){
                    // keep code
                } else {
                    match.parentElement.removeChild(match);
                }

            } else {
                match.parentElement.removeChild(match);
            }
        }
    }""")
            
        elif awaydd.value == "Away 1.3 - 1.5":
            print(awaydd.value)
            selection = awaydd.value + " Finish"
            driver.execute_script("""matches = document.getElementsByClassName("event__match event__match--twoLine");
    for(i=matches.length -1; i>=0; i--){
        match = matches[i];
        if(match.children.length > 9){
            if(match.children[9] && match.children[10] && match.children[11]){
                odd_1 = parseFloat(match.children[9].textContent);
                odd_x = parseFloat(match.children[10].textContent);
                odd_2 = parseFloat(match.children[11].textContent);

                if((odd_1 >= 2.50 && odd_1 <= 50.0) && (odd_x >= 2.5 && odd_x <= 15.0) && (odd_2 >= 1.3 && odd_2 <= 1.5)){
                    // keep code
                } else {
                    match.parentElement.removeChild(match);
                }

            } else {
                match.parentElement.removeChild(match);
            }
        }
    }""")
        
def home_filter_finish(e):
    global driver, selection
        
    if driver:
        if homedd.value == "Home 2.0 - 2.5":
            print(homedd.value)
            selection = homedd.value + " Finish"
            driver.execute_script("""""")
            
        elif homedd.value == "Home 1.5 - 2.0":
            print(homedd.value)
            selection = homedd.value + " Finish"
            driver.execute_script("""""")
            
        elif homedd.value == "Home 1.3 - 1.5":
            print(homedd.value)
            selection = homedd.value + " Finish"
            driver.execute_script("""""")
        
def home_filter_live(e):
    global driver, selection
        
    if driver:
        if homedd.value == "Home 2.0 - 2.5":
            print(homedd.value)
            selection = homedd.value + " Live"
            driver.execute_script("""""")
            
        elif homedd.value == "Home 1.5 - 2.0":
            print(homedd.value)
            selection = homedd.value +" Live"
            driver.execute_script("""""")
            
        elif homedd.value == "Home 1.3 - 1.5":
            print(homedd.value)
            selection = homedd.value +" Live"
            driver.execute_script("""""")
        
def away_filter_live(e):
    global driver, selection
        
    if driver:
        if awaydd.value == "Away 2.0 - 2.5":
            print(awaydd.value)
            selection = awaydd.value + " Live"
            driver.execute_script("""var matches = document.getElementsByClassName("event__match event__match--twoLine")
        for(i = matches.length-1; i>=0; i--){
        match = matches[i];
        if(match.className.includes("event__match--scheduled")){
            if((match.querySelector('div.odds__odd.event__odd--odd1').innerText) && (match.querySelector('div.odds__odd.event__odd--odd2').innerText) && (match.querySelector('div.odds__odd.event__odd--odd3').innerText)){
                //
                var odd_1 = match.querySelector('div.odds__odd.event__odd--odd1').innerText;
                var odd_x = match.querySelector('div.odds__odd.event__odd--odd2').innerText;
                var odd_2 = match.querySelector('div.odds__odd.event__odd--odd3').innerText;

                if((odd_1 >= 2.50 && odd_1 <= 8.00) &&
                (odd_x >= 2.50 && odd_x <= 5.50) && 
                (odd_2 >= 2.00 && odd_2 < 2.50)){
                } else {
                    match.parentElement.removeChild(match);
                }
            } else {
                match.parentElement.removeChild(match);
            }
        } else {
            match.parentElement.removeChild(match);
        }
    }""")
        elif awaydd.value == "Away 1.5 - 2.0":
            print(awaydd.value)
            selection = awaydd.value +" Live"
            driver.execute_script("""var matches = document.getElementsByClassName("event__match event__match--twoLine")

    for(i = matches.length-1; i>=0; i--){
        match = matches[i];
        if(match.className.includes("event__match--scheduled")){
            if((match.querySelector('div.odds__odd.event__odd--odd1').innerText) && (match.querySelector('div.odds__odd.event__odd--odd2').innerText) && (match.querySelector('div.odds__odd.event__odd--odd3').innerText)){
                //
                var odd_1 = match.querySelector('div.odds__odd.event__odd--odd1').innerText;
                var odd_x = match.querySelector('div.odds__odd.event__odd--odd2').innerText;
                var odd_2 = match.querySelector('div.odds__odd.event__odd--odd3').innerText;

                if((odd_1 >= 2.00 && odd_1 <= 7.50) &&
                (odd_x >= 2.00 && odd_x <= 5.50) && 
                (odd_2 >= 1.50 && odd_2 < 2.00)){
                } else {
                    match.parentElement.removeChild(match);
                }
            } else {
                match.parentElement.removeChild(match);
            }
        } else {
            match.parentElement.removeChild(match);
        }
    }""")
            
        elif awaydd.value == "Away 1.3 - 1.5":
            print(awaydd.value)
            selection = awaydd.value +" Live"
            driver.execute_script("""var matches = document.getElementsByClassName("event__match event__match--twoLine")

    for(i = matches.length-1; i>=0; i--){
        match = matches[i];
        if(match.className.includes("event__match--scheduled")){
            if((match.querySelector('div.odds__odd.event__odd--odd1').innerText) && (match.querySelector('div.odds__odd.event__odd--odd2').innerText) && (match.querySelector('div.odds__odd.event__odd--odd3').innerText)){
                //
                var odd_1 = match.querySelector('div.odds__odd.event__odd--odd1').innerText;
                var odd_x = match.querySelector('div.odds__odd.event__odd--odd2').innerText;
                var odd_2 = match.querySelector('div.odds__odd.event__odd--odd3').innerText;

                if((odd_1 >= 2.50 && odd_1 <= 50.00) &&
                (odd_x >= 2.50 && odd_x <= 10.50) && 
                (odd_2 >= 1.30 && odd_2 < 1.50)){
                } else {
                    match.parentElement.removeChild(match);
                }
            } else {
                match.parentElement.removeChild(match);
            }
        } else {
            match.parentElement.removeChild(match);
        }
    }""")
      
def remove_headers(e):
    global drive
    
    if driver:
        driver.execute_script("""const allsoccer = document.getElementsByClassName("sportName soccer")[0].children;
    for (let i = allsoccer.length - 1; i >= 0; i--) {
        const current = allsoccer[i];
        const prev = allsoccer[i - 1];
        
        if (prev) {
            const condition1 = prev.className.includes("wclLeagueHeader") && current.className.includes("wclLeagueHeader")
            // const condition2 = prev.className.includes("event__match") && current.className.includes("event__match")

            if(condition1) {
                prev.parentElement.removeChild(prev)
            }
    }
    }""")
      
def remove_titles(e):
    global driver
    
    if driver:
        driver.execute_script("""bar = document.getElementsByClassName("wclLeagueHeader");
    for(i=bar.length-1; i>=0 ; i--){
        bar[i].parentElement.removeChild(bar[i]);
    }""")

def refresh_page(e):
    global driver
    
    if driver:
        driver.refresh()
        
        sleep(2)
        driver.get(global_url)
        
        auto_run_btn.disabled = False
        auto_run_btn.update()
            
def firefox_close(e):
    global driver, active
    
    if driver:
        active = False
        driver.quit()
        driver = None
        open.disabled = False
        open.update()


def main(page: ft.Page):
    page.title = "SoccerProV1"
    page.window.left = 1000
    page.window.top = 200
    page.window.always_on_top = True
    # page.window.title_bar_hidden = True
    # page.window.frameless = True
    page.bgcolor = ft.colors.TRANSPARENT
    page.window.width = 320
    page.window.height = 600
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    
    page.theme = ft.Theme(
            color_scheme=ft.ColorScheme(
                primary=ft.colors.GREEN,
                primary_container=ft.colors.GREEN_200
                # ...
            ),
        )
    
    global awaydd, homedd, selected_files, open, auto_run_btn
    
    open = ft.ElevatedButton("Open", icon="OPEN_IN_BROWSER", on_click=firefox_launch, width=110)
    close = ft.ElevatedButton("Close", icon="EXIT_TO_APP", on_click=firefox_close, width=110)
    odds = ft.ElevatedButton("Odds", on_click=odds_click, width=100)
    removeTitles = ft.ElevatedButton("Titles", on_click=remove_titles, width=100)
    removeHeaders = ft.ElevatedButton("Header", on_click=remove_headers, width=100)
    
    awaydd = ft.Dropdown(
        value="Away 2.0 - 2.5",
        width=150,
        options=[
            ft.dropdown.Option("Away 2.0 - 2.5"),
            ft.dropdown.Option("Away 1.5 - 2.0"),
            ft.dropdown.Option("Away 1.3 - 1.5"),
        ],
    )
    
    homedd = ft.Dropdown(
        value="Home 2.0 - 2.5",
        width=150,
        options=[
            ft.dropdown.Option("Home 2.0 - 2.5"),
            ft.dropdown.Option("Home 1.5 - 2.0"),
            ft.dropdown.Option("Home 1.3 - 1.5"),
        ],
    )
    
    finishAway = ft.IconButton(
                icon=ft.icons.SPORTS_BASEBALL,
                icon_color="red400",
                icon_size=30,
                tooltip="Finish",
                on_click=away_filter_finish,
            )
    
    finishHome = ft.IconButton(
            icon=ft.icons.SPORTS_BASEBALL,
            icon_color="red400",
            icon_size=30,
            tooltip="Finish",
            on_click=home_filter_finish,
        )
    
    liveAway = ft.IconButton(
                    icon=ft.icons.SPORTS_BASEBALL,
                    icon_color="green400",
                    icon_size=40,
                    tooltip="Live",
                    on_click=away_filter_live,
                )
    
    liveHome = ft.IconButton(
                icon=ft.icons.SPORTS_BASEBALL,
                icon_color="green400",
                icon_size=40,
                tooltip="Live",
                on_click=home_filter_live,
            )
    
    refresh = ft.IconButton(
        icon=ft.icons.REFRESH,
        icon_color="orange300",
        icon_size=30,
        tooltip="Refresh",
        on_click=refresh_page,
    )
    
    all_matches_url_btn = ft.ElevatedButton("Matches", icon="ADD_HOME_WORK", on_click=get_all_matches_url, )
    open_new_tab_btn = ft.ElevatedButton("new_tab", icon="BACKUP_TABLE", on_click=open_new_tab)
    odds_btn = ft.ElevatedButton("odds", icon="TABLE_BAR_OUTLINED", on_click=get_odds)
    last_event_btn = ft.ElevatedButton("last match", icon="PIVOT_TABLE_CHART_SHARP", on_click=get_last_match)
    table_btn = ft.ElevatedButton("table match", icon="TABLE_CHART_SHARP", on_click=details_from_table)
    auto_run_btn = ft.FilledTonalButton("Run data", icon="AUTO_FIX_HIGH_OUTLINED", on_click=auto_run_data)
    
    content1 = ft.Container(
                    ft.Column([
                        ft.Row([open, close], alignment=ft.MainAxisAlignment.CENTER,),
                        ft.Row([odds, refresh], alignment=ft.MainAxisAlignment.CENTER,),
                        ft.Row([removeHeaders, removeTitles], alignment=ft.MainAxisAlignment.CENTER,),
                        ft.Card(content=ft.Container(
                            ft.Column([
                                ft.Row([awaydd, finishAway, liveAway], alignment=ft.MainAxisAlignment.CENTER,),
                                ft.Row([homedd, finishHome, liveHome], alignment=ft.MainAxisAlignment.CENTER,),
                                
                            ]),
                            padding=10,
                        )),
                        ft.Row([all_matches_url_btn, open_new_tab_btn], alignment=ft.MainAxisAlignment.CENTER,),
                        ft.Row([odds_btn, last_event_btn, ], alignment=ft.MainAxisAlignment.CENTER,),
                        ft.Row([table_btn, auto_run_btn], alignment=ft.MainAxisAlignment.CENTER,),
                    ], alignment=ft.MainAxisAlignment.CENTER,)
                )
    
    
    # Tab 2
    
    global scores_broswer_btn

    file_picker = ft.FilePicker(on_result=open_file)
    page.overlay.append(file_picker)
    
    file_btn = ft.OutlinedButton(text="Choose XLSX", icon="UPLOAD_FILE", on_click=lambda _: file_picker.pick_files(allowed_extensions=["xlsx"]))
    sftr = ft.Text("Search for the result")
    
    scores_broswer_btn = ft.ElevatedButton("open broswer", icon="TABLE_BAR_OUTLINED", on_click=get_scores)
    
    scores_btn = ft.ElevatedButton("visit url", icon="TABLE_BAR_OUTLINED", on_click=get_all_scores)
    
    analysis_btn = ft.ElevatedButton("Analysis", icon="SPORTS_BASEBALL", icon_color="yellow500", on_click=get_analysis)
    
    selected_files = ft.Text()
    
    content2 = ft.Container(
                    ft.Column([
                        ft.Row([sftr], alignment=ft.MainAxisAlignment.CENTER,),
                        ft.Row([file_btn], alignment=ft.MainAxisAlignment.CENTER,),
                        ft.Row([selected_files], alignment=ft.MainAxisAlignment.CENTER,),
                        ft.Row([scores_broswer_btn, scores_btn], alignment=ft.MainAxisAlignment.CENTER,),
                        ft.Row([analysis_btn], alignment=ft.MainAxisAlignment.CENTER,),
                    ], alignment=ft.MainAxisAlignment.CENTER,)
                )
    
    
    t = ft.Tabs(
        selected_index=1,
        animation_duration=300,
        tabs=[
            ft.Tab(
                text="Home",
                content=content1,
            ),
            ft.Tab(
                text="Final Result",
                icon=ft.icons.SEARCH,
                content=content2,
            ),    
            ft.Tab(
                text="Settings",
                tab_content=ft.Icon(ft.icons.SETTINGS),
                content=ft.Text("This is Tab 3"),
            ),
        ],
        expand=1,
    )
    
    page.add(t)
    
ft.app(main)
