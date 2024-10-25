import openpyxl
from openpyxl.styles import PatternFill
import os

workbook = openpyxl.load_workbook("demo.xlsx")
sheet = workbook.active

# Specify the range of cells to fill (e.g., cells A1 to B5)
start_row = 1
start_col = 1
end_row = 5
end_col = 5

# fill = PatternFill(fill_type=None,
#                 start_color='FFFFFFFF',
#                 end_color='FF000000')

# Fill the range with numbers
for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.value = row * col + 2
        cell.fill = PatternFill("solid", fgColor="009999FF")

# Save the updated workbook
workbook.save("demo.xlsx")
os.startfile("demo.xlsx")


    # last_row = sheet.max_row
    
    # while last_row > 1:
    #     row = sheet[last_row]
    #     if any(cell.value for cell in row):
    #         break  # Found a row with content
    #     last_row -= 1

    # # Access the last row with content
    # last_row_data = sheet[last_row]
    # for cell in last_row_data:
    #     print(cell.value)


# import openpyxl

# workbook = openpyxl.load_workbook("your_file.xlsx")
# sheet = workbook.active

# column_index = 2  # Assuming column B

# for row in sheet.iter_rows(min_row=2, max_row=kln jsheet.max_row, values_only=True):
#     cell_value = row[column_index - 1]
#     new_value = cell_value * 2  # Example: Double the value
#     sheet[column_letter][row_index] = new_value

# workbook.save("your_file.xlsx")



# import openpyxl
# import os


# path = os.path.join("soccerpro", "output_data", "1.5-2.0.xlsx")

# workbook = openpyxl.load_workbook(path)
# sheet = workbook.active

# column_letter = "B" # Assuming URLs are in column B

# urls = [cell.value for cell in sheet[column_letter]]

# for url in urls:
#     print(url)



# # import xlsxwriter module 
# import os
# import xlsxwriter 
 
# # Workbook() takes one, non-optional, argument 
# # which is the filename that we want to create. 
# workbook = xlsxwriter.Workbook('sample.xlsx') 
 
# # The workbook object is then used to add new 
# # worksheet via the add_worksheet() method. 
# worksheet = workbook.add_worksheet() 
 
# # Use the worksheet object to write 
# # data via the write() method. 
# worksheet.write('A1', 'Hello..') 
# worksheet.write('B1', 'Geeks') 
# worksheet.write('C1', 'For') 
# worksheet.write('D1', 'Geeks') 
 
# # Finally, close the Excel file 
# # via the close() method. 
# workbook.close() 
# os.startfile("sample.xlsx")